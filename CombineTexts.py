#
#   CombineText.py
#   PythonUtilies
#
#   Created by Kelly Chui on 2022/09/30
#

import os
from openpyxl import Workbook

directory = "C:/texts/"
files = os.listdir(directory)
excelFile = Workbook()
write_ws = excelFile.create_sheet(index=1, title='result')

i = 1
for filename in files:
    if ".txt" not in filename:
        continue
    j = 1
    file = open(directory + filename)
    contents = []
    for line in file:
        if "=" in line:
            splits = line.split("=")
            write_ws.cell(j, i, splits[1].rstrip("\n"))
        j = j + 1
    i = i + 1
excelFile.save("C:/result/result.xlsx")