#
#   CombineText.py
#   PythonUtilties
#
#   Created by Kelly Chui on 2022/09/30
#

import os
import webbrowser
from openpyxl import Workbook
import natsort
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

def selectLoadFolder():
    global loadPath
    loadPath = filedialog.askdirectory(initialdir = "/", title = "Open text data")
    if loadPath == "":
        tk.messagebox.showwarning("Warning", "It's not a folder.")
    else:
        res = os.listdir(loadPath)
        print(res)
        if len(res) == 0:
            messagebox.showwarning("Warning", "There are no files inside the folder.")
        else:
            loadPathLabel.config(text=loadPath)
            for file in res:
                print(loadPath + "/" + file)

def selectSaveFolder():
    global savePath
    savePath = filedialog.askdirectory(initialdir = "/", title = "Select a save directory")
    savePathLabel.config(text=savePath)

def makeExcelFile(savePath):
    resultFile = Workbook()
    resultFile.remove(resultFile["Sheet"])
    sheet = resultFile.create_sheet(index = 1, title = "result")

    if loadPath == "" or savePath == "":
        messagebox.showwarning("Warning", "Directories must be selected")
        return
    files = os.listdir(loadPath)
    files = list(filter(lambda x: ".txt" in x, files))
    files = natsort.natsorted(files)
    print(files)
    row = 1
    for fileName in files:
        column = 1
        file = open(loadPath + "/" + fileName)
        for line in file:
            if "=" in line:
                splits = line.split("=")
                try: 
                    cellItem = float(splits[1].rstrip("\n"))
                except:
                    cellItem = "failed"
                sheet.cell(column, row, cellItem)
            column += 1
        row += 1

    resultFile.save(savePath + "/" + "result" + ".xlsx")
    print("end")
    messagebox.showinfo("Complete", "Complete")

def openExplorer():
    os.system(f"explorer {os.path.realpath(savePath)}")

def openURL():
    webbrowser.open("https://github.com/Kelly-Chui/PythonUtilities")

def openpyxlURL():
    webbrowser.open("https://openpyxl.readthedocs.io/en/stable/")

def openNatsortURL():
    webbrowser.open("https://github.com/SethMMorton/natsort")

def showAppinfo():
    appInfoView = tk.Toplevel()
    appInfoView.title("App infomation")
    appInfoView.geometry("200x200")
    appInfoView.resizable(False, False)
    maker = tk.Label(appInfoView, text = "Made by Kelly Chui")
    githubURL = tk.Button(appInfoView, text = "contact(Website)", command = openURL)
    openSourceLabel = tk.Label(appInfoView, text = "Opensource License")
    openSourceInfo1 = tk.Button(appInfoView, text = "natsort", comman = openNatsortURL)
    openSourceInfo2 = tk.Button(appInfoView, text = "openpyxl", command = openpyxlURL)

    maker.pack(side="top")
    githubURL.pack(side="top")

    openSourceInfo1.pack(side="bottom")
    openSourceInfo2.pack(side="bottom")
    openSourceLabel.pack(side="bottom")

rootView = tk.Tk()
rootView.configure(bg = "white")
rootView.title("TextToExcel")
rootView.geometry("600x400")
rootView.resizable(False, False)

file = tk.Frame(rootView)
file.pack(fill = "x", padx = 5, pady = 5)

loadPath = ""
savePath = ""
textList = []

loadLabel = tk.Label(rootView, bg = "white", height = 10)
loadLabel.pack(side = "top", pady = 10)
saveLabel = tk.Label(rootView, bg = "white", height = 10)
saveLabel.pack(side = "top", pady = 10)
runLabel = tk.Label(rootView, bg = "white")
runLabel.pack()
infoLabel = tk.Label(rootView, bg = "white", height = 20)
infoLabel.pack(side = "bottom")

loadPathLabel = tk.Label(loadLabel, text = "loadPath", width = "30")
savePathLabel = tk.Label(saveLabel, text = "savePath", width = "30")
loadButton = tk.Button(loadLabel, text = "Load", bg = "white", width = 5, command = selectLoadFolder)
saveButton = tk.Button(saveLabel, text = "Save", width = 5, bg = "white", command = selectSaveFolder)
proceedButton = tk.Button(runLabel, text = "Run", fg = "white", bg = "green", width = 10, pady = 5, command = lambda: makeExcelFile(savePath))
openSaveButton = tk.Button(runLabel, text = "Show results in Explorer", bg = "white", fg = "blue", width = 20, command = openExplorer)
infoButton = tk.Button(infoLabel, text = "App info", bg = "white", fg = "green", command = showAppinfo)

loadButton.pack(side = "right", padx = 10)
loadPathLabel.pack(side = "right")
saveButton.pack(side = "right", padx = 10)
savePathLabel.pack(side = "right")
proceedButton.pack(side = "top")
openSaveButton.pack(side = "top")
infoButton.pack()
rootView.mainloop()
